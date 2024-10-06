#import os
from openpyxl import Workbook
from openpyxl import load_workbook

def WriteReportCertificado(DicInputs, DicPrecio, avg_Precio, 
                AlcobasToPlot, IPVU_Delta_List, avg_M2Construidos):

    # grab the active worksheet
    wb = load_workbook('Template.xlsx')
    ws = wb.active

    # Cheecking the type of ID 
    if DicInputs['TipoDeDocumento'] == 'Cédula de Ciudadanía':
        DicInputs['TipoDeDocumento'] = 'CC'
    elif DicInputs['TipoDeDocumento'] == 'Cédula de Extranjería':
        DicInputs['TipoDeDocumento'] = 'CE'
    elif DicInputs['TipoDeDocumento'] == 'Pasaporte':
        DicInputs['TipoDeDocumento'] = 'P'
    
    # Assigning data to cells

    ws['Q11'] = DicInputs['Nombre'] + ' ' + DicInputs['Apellidos']
    # ws[''] = DicInputs['Email']
    ws['AD58'] = DicInputs['Alcobas']
    ws['AQ58'] = DicInputs['Banos']
    ws['AQ60'] = DicInputs['BanosSociales']
    ws['AQ62'] = DicInputs['BanosPrivados']
    ws['C46'] = DicInputs['Celular']


    ws['L62'] = DicInputs['Niveles']
    ws['DG58'] = DicInputs['Administracion']
    ws['AB60'] = DicInputs['AreaTerraza']
    ws['Y54'] = DicInputs['AnoConstruccion']
    #ws['AD62'] = DicInputs['CantidadParqueaderos']
    ws['DT61'] = DicInputs['CantidadParqueaderos']
    ws['EF22'] = DicInputs['Estrato']
    ws['AY89'] = DicInputs['M2Construidos']
    ws['L58'] = DicInputs['Piso']
    ws['DG60'] = DicInputs['Predial']
    ws['AZ13'] =  DicInputs['TipoDeVia']+ \
                  ' ' + \
                  DicInputs['Numero1']+ \
                  ' # ' + \
                  DicInputs['Numero2']+ \
                  ' - ' + \
                  DicInputs['Numero3']+ \
                  ' - Apto ' + \
                  DicInputs['NumeroApartamento']
    ws['J15'] = DicInputs['Ciudad']
    ws['BF15']= DicInputs['Departamento']
    # ws[''] = DicInputs['TipoDeInmueble']
    ws['FM13'] = DicInputs['NombreEdificio']
    ws['DW13'] = DicInputs['Barrio']
    ws['CV15'] = DicInputs['TipoDeSector']
    ws['Z50'] = DicInputs['PisosEdificio']
    ws['Z52'] = DicInputs['SotanosEdificio']
    ws['L60'] = DicInputs['Salas']
    ws['BJ58'] = DicInputs['Comedores']
    ws['BJ60'] = DicInputs['Estudios']
    ws['BJ62'] = DicInputs['Balcones']
    ws['EW58'] = DicInputs['ParqueaderosCubiertos']
    ws['EW60'] = DicInputs['ParqueaderosDescubiertos']
    ws['FQ58'] = DicInputs['ParqueaderosSencillos']
    ws['FQ60'] = DicInputs['ParqueaderosDobles']
    ws['CE11'] = DicInputs['TipoDeDocumento']+ ' ' + DicInputs['NumeroDeDocumento']
    ws['BN58'] = ', '.join(DicInputs['DetallesApartamento'])
    ws['BG75'] = ', '.join(DicInputs['DetallesEdificio'])
    ws['D22'] = ', '.join(DicInputs['DetallesSector'])
    #ws['EB58'] = ', '.join(DicInputs['DetallesGaraje'])
    ws['B89'] = 'Apartamento ' + DicInputs['NumeroApartamento'] + ' - ' + DicInputs['NombreEdificio']    
    ws['CV89'] = int(round(DicPrecio['PrecioM2'][0], -5))




    wb.save("ReporteCertificado.xlsx")
    
    
    return None 
