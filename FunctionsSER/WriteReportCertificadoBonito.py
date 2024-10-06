#import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import numpy as np

def WriteReportCertificadoBonito(DicInputs, df_apartamentos_neigh, PrecioByLaw, DicLocation):

    # Cheecking the type of ID 
    if DicInputs['TipoDeDocumento'] == 'Cédula de Ciudadanía':
        DicInputs['TipoDeDocumento'] = 'CC'
    elif DicInputs['TipoDeDocumento'] == 'Cédula de Extranjería':
        DicInputs['TipoDeDocumento'] = 'CE'
    elif DicInputs['TipoDeDocumento'] == 'Pasaporte':
        DicInputs['TipoDeDocumento'] = 'P'
 
    # grab the active worksheet
    wb = load_workbook('Templates/TemplateBonito_2022_12_21.xlsx')
    ws = wb["mercado"]

   
    # Assigning data to cells
    ws['C10'] = DicInputs['NombrePerito']

    ws['C44'] = DicInputs['Nombre'] + ' ' + DicInputs['Apellidos']
    ws['C45'] = DicInputs['TipoDeDocumento']+ ' ' + DicInputs['NumeroDeDocumento'] 
    ws['C46'] = DicInputs['Celular']
    ws['C48'] = DicInputs['Email']

    ws['C54'] = str(DicLocation['Latitud']) + ', ' +str(DicLocation['Longitud'])
    ws['C56'] = DicInputs['TipoDeSector']
    ws['C57'] =  DicInputs['TipoDeVia']+ \
                  ' ' + \
                  DicInputs['Numero1']+ \
                  ' # ' + \
                  DicInputs['Numero2']+ \
                  ' - ' + \
                  DicInputs['Numero3']+ \
                  ' - Apto ' + \
                  DicInputs['NumeroApartamento']
    ws['C58'] = DicInputs['NombreEdificio']
    ws['C59'] = DicInputs['Barrio']
    ws['C60'] = DicInputs['Ciudad']
    ws['C61']= DicInputs['Departamento']

    ws['C68'] = DicInputs['Administracion']
    ws['C69'] = DicInputs['Predial']
    
    ws['C104'] = DicInputs['Estrato']

    ws['F132'] = DicInputs['M2Construidos']

    #writing the comparative table and the images
    excel_row = 343
    image_anchors = ['A320','E320','A334','E334','A347','E347','A360', 'A375', 'E375', 'A388']
    for  row in range(10):
        #writing table    
        ws['C'+str(excel_row)] = df_apartamentos_neigh.iloc[row]['SitioWeb'] #'redinmobiliariamls.com'
        ws['D'+str(excel_row)] = df_apartamentos_neigh.iloc[row]['Barrio'] 

        if not np.isnan(df_apartamentos_neigh.iloc[row]['Estrato']):
            ws['F'+str(excel_row)] = int(df_apartamentos_neigh.iloc[row]['Estrato'])
        ws['G'+str(excel_row)] = df_apartamentos_neigh.iloc[row]['Area']
        ws['I'+str(excel_row)] = df_apartamentos_neigh.iloc[row]['Precio']
#        ws['K'+str(excel_row)] = int(2022 - df_apartamentos_neigh.iloc[row]['Año de construcción'])
        ws['D'+str(excel_row+17)] = df_apartamentos_neigh.iloc[row]['Parqueaderos']
        excel_row += 1

        #pasting Images
        ws = wb['Formato']
        img = Image(df_apartamentos_neigh.iloc[row]['ScreenshotPath'])
        img.anchor = image_anchors[row]
        img.height = 384/1.7
        img.width  = 683/1.7
        ws.add_image(img)
        ws = wb['mercado']



    #pasting the map
    ws = wb['Formato']
    img = Image('Mapa.png')
    img.anchor = 'B75' #image_anchors[row]
#    img.height = 384/1.7
#    img.width  = 683/1.7
    ws.add_image(img)
    ws = wb['mercado']
    


    if "Acueducto" in DicInputs['DetallesApartamento']:
        ws['D74'] = 'SI'
    else:
        ws['D74'] = 'NO'
    if "Alcantarillado" in DicInputs['DetallesApartamento']:
        ws['D75'] = 'SI'
    else:
        ws['D75'] = 'NO'
    if "Energía Eléctrica" in DicInputs['DetallesApartamento']:
        ws['D76'] = 'SI'
    else:
        ws['D76'] = 'NO'
    if "Gas Natural" in DicInputs['DetallesApartamento']:
        ws['D77'] = 'SI'
    else:
        ws['D77'] = 'NO'
    if "Telefonía" in DicInputs['DetallesApartamento']:
        ws['D78'] = 'SI'
    else:
        ws['D78'] = 'NO'
    if "Internet" in DicInputs['DetallesApartamento']:
        ws['D79'] = 'SI'
    else:
        ws['D79'] = 'NO'


    if "Acueducto" in DicInputs['DetallesSector']:
        ws['C74'] = 'SI'
    else:
        ws['C74'] = 'NO'
    if "Alcantarillado" in DicInputs['DetallesSector']:
        ws['C75'] = 'SI'
    else:
        ws['C75'] = 'NO'
    if "Energía Eléctrica" in DicInputs['DetallesSector']:
        ws['C76'] = 'SI'
    else:
        ws['C76'] = 'NO'
    if "Gas Natural" in DicInputs['DetallesSector']:
        ws['C77'] = 'SI'
    else:
        ws['C77'] = 'NO'
    if "Telefonía" in DicInputs['DetallesSector']:
        ws['C78'] = 'SI'
    else:
        ws['C78'] = 'NO'
    if "Internet" in DicInputs['DetallesSector']:
        ws['C79'] = 'SI'
    else:
        ws['C79'] = 'NO'

    if "Parques Cercanos" in DicInputs['DetallesSector']:
        ws['C97'] = 'SI'
    else:
        ws['C97'] = 'NO'
    if "Parqueaderos" in DicInputs['DetallesSector']:
        ws['C98'] = 'SI'
    else:
        ws['C98'] = 'NO'
    if "Alumbrado Público" in DicInputs['DetallesSector']:
        ws['C99'] = 'SI'
    else:
        ws['C99'] = 'NO'
    if "Zonas Verdes" in DicInputs['DetallesSector']:
        ws['C100'] = 'SI'
    else:
        ws['C100'] = 'NO'
    if "Arborización" in DicInputs['DetallesSector']:
        ws['E97'] = 'SI'
    else:
        ws['E97'] = 'NO'
    if "Alamedas" in DicInputs['DetallesSector']:
        ws['E98'] = 'SI'
    else:
        ws['E98'] = 'NO'
    if "Ciclorutas Cercanas" in DicInputs['DetallesSector']:
        ws['E99'] = 'SI'
    else:
        ws['E99'] = 'NO'
    if "Comercio" in DicInputs['DetallesSector']:
        ws['E100'] = 'SI'
    else:
        ws['E100'] = 'NO'
    if "Recreación" in DicInputs['DetallesSector']:
        ws['G97'] = 'SI'
    else:
        ws['G97'] = 'NO'
    if "Servicios de Salud" in DicInputs['DetallesSector']:
        ws['G98'] = 'SI'
    else:
        ws['G98'] = 'NO'
    if "Servicios Bancarios" in DicInputs['DetallesSector']:
        ws['G99'] = 'SI'
    else:
        ws['G99'] = 'NO'
    if "Servicios Escolares" in DicInputs['DetallesSector']:
        ws['G100'] = 'SI'
    else:
        ws['G100'] = 'NO'

    ws['C155'] = DicInputs['PisosEdificio']
    ws['C156'] = DicInputs['SotanosEdificio']
    ws['C157'] = DicInputs['AnoConstruccion']
    ws['C159'] = DicInputs['Niveles']



    ws['C177'] = DicInputs['Salas']
    ws['C178'] = DicInputs['Comedores']
    ws['C179'] = DicInputs['Estudios']
    ws['C180'] = DicInputs['Estares']
    ws['C181'] = DicInputs['Alcobas']
    ws['C182'] = DicInputs['BanosPrivados']
    ws['C183'] = DicInputs['BanosSociales']
    ws['C184'] = DicInputs['Cocinas']
    ws['C185'] = DicInputs['AlcobasDeServicio']
    ws['C186'] = DicInputs['BanosDeServicio']
    ws['C187'] = DicInputs['Patios']
    ws['C188'] = DicInputs['Balcones']
    ws['C189'] = DicInputs['Terrazas']
    ws['C191'] = DicInputs['CuartosUtiles']

    if "Arborización" in DicInputs['DetallesSector']:
        ws['F198'] = 'SI'
    else:
        ws['F198'] = 'NO'
    if "Parques Cercanos" in DicInputs['DetallesSector']:
        ws['F199'] = 'SI'
    else:
        ws['F199'] = 'NO'
    if "Zonas Verdes" in DicInputs['DetallesSector']:
        ws['F200'] = 'SI'
    else:
        ws['F200'] = 'NO'
    

    if "Portería" in DicInputs['DetallesEdificio']:
        ws['C224'] = 'SI'
    else:
        ws['C224'] = 'NO'
    if "Citófono" in DicInputs['DetallesEdificio']:
        ws['C225'] = 'SI'
    else:
        ws['C225'] = 'NO'
    if "Parqueadero Visitantes" in DicInputs['DetallesEdificio']:
        ws['C226'] = 'SI'
    else:
        ws['C226'] = 'NO'
    if "Bicicletero" in DicInputs['DetallesEdificio']:
        ws['C227'] = 'SI'
    else:
        ws['C227'] = 'NO'
    if "Club House" in DicInputs['DetallesEdificio']:
        ws['C228'] = 'SI'
    else:
        ws['C228'] = 'NO'
    if "Salón Social" in DicInputs['DetallesEdificio']:
        ws['C229'] = 'SI'
    else:
        ws['C229'] = 'NO'
    if "Piscina" in DicInputs['DetallesEdificio']:
        ws['C230'] = 'SI'
    else:
        ws['C230'] = 'NO'
    if "Juegos de Niños" in DicInputs['DetallesEdificio']:
        ws['C231'] = 'SI'
    else:
        ws['C231'] = 'NO'
    if "Zonas Verdes" in DicInputs['DetallesEdificio']:
        ws['C232'] = 'SI'
    else:
        ws['C232'] = 'NO'
    if "Salón de Juegos" in DicInputs['DetallesEdificio']:
        ws['C233'] = 'SI'
    else:
        ws['C233'] = 'NO'

    if "Cancha Múltiple" in DicInputs['DetallesEdificio']:
        ws['F224'] = 'SI'
    else:
        ws['F224'] = 'NO'
    if "Gimnasio" in DicInputs['DetallesEdificio']:
        ws['F225'] = 'SI'
    else:
        ws['F225'] = 'NO'
    if "Cancha de Squash" in DicInputs['DetallesEdificio']:
        ws['F226'] = 'SI'
    else:
        ws['F226'] = 'NO'
    if "Golfito" in DicInputs['DetallesEdificio']:
        ws['F227'] = 'SI'
    else:
        ws['F227'] = 'NO'
    if "Aire Acondicionado Central" in DicInputs['DetallesEdificio']:
        ws['F228'] = 'SI'
    else:
        ws['F228'] = 'NO'
    if "Zonas Húmedas" in DicInputs['DetallesEdificio']:
        ws['F229'] = 'SI'
    else:
        ws['F229'] = 'NO'
    if "Ascensor" in DicInputs['DetallesEdificio']:
        ws['F230'] = 'SI'
    else:
        ws['F230'] = 'NO'
    if "Sendero" in DicInputs['DetallesEdificio']:
        ws['F231'] = 'SI'
    else:
        ws['F231'] = 'NO'

    if "Planta Eléctrica" in DicInputs['DetallesEdificio']:
        ws['I224'] = 'SI'
    else:
        ws['I224'] = 'NO'
    if "Tanque de Agua" in DicInputs['DetallesEdificio']:
        ws['I225'] = 'SI'
    else:
        ws['I225'] = 'NO'
    if "Bomba" in DicInputs['DetallesEdificio']:
        ws['I226'] = 'SI'
    else:
        ws['I226'] = 'NO'
    if "Sistema de Presión" in DicInputs['DetallesEdificio']:
        ws['I227'] = 'SI'
    else:
        ws['I227'] = 'NO'
    if "Shut de Basura" in DicInputs['DetallesEdificio']:
        ws['I228'] = 'SI'
    else:
        ws['I228'] = 'NO'
    if "Bomba Eyectora" in DicInputs['DetallesEdificio']:
        ws['I229'] = 'SI'
    else:
        ws['I229'] = 'NO'


    ws['B242'] = DicInputs['AnalisisInmueble']
    ws['B256'] = DicInputs['ActividadEdificadora']   
    ws['B268'] = DicInputs['OfertaDemanda']           
    ws['B286'] = DicInputs['ObservacionesSector']   
    ws['B293'] = DicInputs['DescInmuebleAcabados'] 
    ws['B300'] = DicInputs['ObsConstEstructura']  
    ws['B307'] = DicInputs['ObsLiquidación']          
    ws['B315'] = DicInputs['ObsGenerales']            
    ws['B111'] = DicInputs['PerspValoracion']

    wb.save("ReporteCertificadoBonito.xlsx")
    
    
    return None 
