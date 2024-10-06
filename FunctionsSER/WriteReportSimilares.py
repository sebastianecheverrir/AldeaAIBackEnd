#import os
from openpyxl import Workbook
#from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import numpy as np

def WriteReportSimilares(DicInputs, df_apartamentos_neigh, PrecioByLaw, DicLocation):

#    print(df_apartamentos_neigh.columns, flush=True)

    # grab the active worksheet
    wb = Workbook() 
    ws = wb.active
    excel_row = 27

    #Add Logo
    img = Image('Images/Logos/LogoOriginal.png')
    img.anchor = 'A1' #image_anchors[row]
    img.height = img.height/2.2
    img.width  = img.width/2.2
    ws.add_image(img)


    # Add map with location of the referene property
    img = Image('Mapa.png')
    img.anchor = 'C5' #image_anchors[row]
#    img.height = 384/1.7
#    img.width  = 683/1.7
    ws.add_image(img)




    # write header
    ws['B'+str(excel_row-1)] = 'ID'
    ws['C'+str(excel_row-1)] = 'SitioWeb'
    ws['D'+str(excel_row-1)] = 'Barrio'
    ws['E'+str(excel_row-1)] = 'Estrato'
    ws['F'+str(excel_row-1)] = 'Area [m2]'
    ws['G'+str(excel_row-1)] = 'Precio [COP]'
    ws['H'+str(excel_row-1)] = 'Parqueaderos'
    ws['I'+str(excel_row-1)] = 'Baños'
    ws['J'+str(excel_row-1)] = 'Habitaciones'
    ws['K'+str(excel_row-1)] = 'Ciudad'
    ws['L'+str(excel_row-1)] = 'Precio m2 [COP]'
    ws['M'+str(excel_row-1)] = 'Fecha de Acceso'



    #writing the comparative table and the images
    image_anchors = []

    #cells where the images will be anchored
    for i in range(len(df_apartamentos_neigh)):
        image_anchors.append('B'+str(excel_row + 23+15*i))
#    image_anchors = ['A20','A35', 'A50', 'A75', 'A90', 'A105', 'A120', 'A135', 'A150', 'A165']

    for  row in range(len(df_apartamentos_neigh)):
        #writing table    
        ws['B'+str(excel_row)] = row+1
        ws['C'+str(excel_row)] = df_apartamentos_neigh.iloc[row]['SitioWeb'] #'redinmobiliariamls.com'
        ws['D'+str(excel_row)] = df_apartamentos_neigh.iloc[row]['Barrio'] 

        if not np.isnan(df_apartamentos_neigh.iloc[row]['Estrato']):
            ws['E'+str(excel_row)] = int(df_apartamentos_neigh.iloc[row]['Estrato'])
        ws['F'+str(excel_row)] = df_apartamentos_neigh.iloc[row]['Area']
        ws['G'+str(excel_row)] = df_apartamentos_neigh.iloc[row]['Precio']
#        ws['K'+str(excel_row)] = int(2022 - df_apartamentos_neigh.iloc[row]['Año de construcción'])
        ws['H'+str(excel_row)] = df_apartamentos_neigh.iloc[row]['Parqueaderos']
        ws['I'+str(excel_row)] = df_apartamentos_neigh.iloc[row]['Banos']
        ws['J'+str(excel_row)] = df_apartamentos_neigh.iloc[row]['Habitaciones']
        ws['K'+str(excel_row)] = df_apartamentos_neigh.iloc[row]['Ciudad']
        ws['L'+str(excel_row)] = df_apartamentos_neigh.iloc[row]['Precio']/df_apartamentos_neigh.iloc[row]['Area']
        ws['M'+str(excel_row)] = df_apartamentos_neigh.iloc[row]['ExtractionDate']

        excel_row += 1

        #pasting Images
#        ws = wb['Formato']
        img = Image(df_apartamentos_neigh.iloc[row]['ScreenshotPath'])
        img.anchor = image_anchors[row]
        img.height = 384/1.7
        img.width  = 683/1.7
        ws.add_image(img)
#        ws = wb['mercado']



    wb.save("Reports/ReporteSimilares.xlsx")
    
    
    return None 
